#!/usr/bin/env python3

import logging
import datetime
import io

from tkinter import NORMAL, DISABLED, END, BOTH, X, Y, TOP, BOTTOM, LEFT, RIGHT, ALL, WORD, FIRST, LAST, GROOVE
from tkinter import Tk, Frame, Label, Button, Entry, Toplevel, scrolledtext, Checkbutton, IntVar
from openpyxl import load_workbook
from tkcalendar import DateEntry
from svglib.svglib import svg2rlg
from reportlab.graphics import renderPM, renderPDF, renderSVG

from templating import TEMPLATE, SAMPLE

from dialogues import save_image, save_postscript, get_file_name, export_workbook
from filing import get_path, get_config_data, save_config_data, wipe_config_file, get_log, wipe_log, append_log
from utils import copy_to_clipboard
from templating import create_template, populate_template, to_date_object, reformat_dates
from checker import check_date_field_formats, check_merged_cells, check_header_rows, check_header_rows_exist, check_sheets_exist, check_misspelled_headers
from settings import Settings

from loader import Loader
from cleaner import Cleaner
from processor import Processor
from painter import Painter
from generator import Generator


class App(Tk):
    def __init__(self):
        super(App, self).__init__()

        self.win_x = int(self.winfo_screenwidth() * 0.1)
        self.win_y = int(self.winfo_screenheight() * 0.1)
        self.geometry(f'+{self.win_x}+{self.win_y}')  # w, h, x, y
        self.resizable(False, False)
        self.title("Gantt Page")
        self.wm_iconbitmap(get_path("favicon.ico"))
        self.controls = Controls(self)
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.mainloop()

    def on_close(self):
        wipe_log()
        try:
            self.quit()
        except RuntimeError:
            self.destroy()


class Controls(Frame):
    def __init__(self, parent):
        super(Controls, self).__init__(parent)

        self.pack()
        self.configure(padx=10, pady=5)
        self.columnconfigure(0, weight=1)
        self.columnconfigure(1, weight=1)

        self.parent = parent  # App is the parent
        self.view = None  # for View (i.e. TopLevel) instance, parent of Chart (i.e. Canvas) instance
        self.file_source = None  # for path to user's Excel spreadsheet
        self.settings = get_config_data()
        self.check_count = 0
        self.run_count = 0
        self.show_rows = IntVar()
        self.show_row_nums = IntVar()

        self.v_cmd_1 = (self.register(self.field_validation_1), '%d', '%i', '%P', '%s', '%S', '%v', '%V', '%W')
        self.v_cmd_2 = (self.register(self.field_validation_2), '%d', '%i', '%P', '%s', '%S', '%v', '%V', '%W')
        self.v_cmd_3 = (self.register(self.field_validation_3), '%d', '%i', '%P', '%s', '%S', '%v', '%V', '%W')

        self.lbl_start = Label(self, text="Timescale start:")
        self.lbl_finish = Label(self, text="Timescale finish:")
        self.lbl_width = Label(self, text="Chart width:")
        self.lbl_height = Label(self, text="Chart height:")
        self.lbl_num_rows = Label(self, text="Number of rows:")
        self.lbl_row_height = Label(self, text="Row height:")
        self.lbl_source = Label(self, text="Source file:")

        self.ent_start = DateEntry(self, date_pattern='yyyy/MM/dd', relief="groove")
        self.ent_finish = DateEntry(self, date_pattern='yyyy/MM/dd', relief="groove")

        self.ent_width = Entry(self, relief="groove", validate="key", validatecommand=self.v_cmd_1)
        self.ent_height = Entry(self, relief="groove", validate="key", validatecommand=self.v_cmd_1)
        self.ent_num_rows = Entry(self, relief="groove", validate="key", validatecommand=self.v_cmd_3)
        self.ent_row_height = Entry(self, relief="groove", validate="key", validatecommand=self.v_cmd_3)
        self.ent_filepath = Entry(self, text="", relief="groove", validate="key", validatecommand=self.v_cmd_2)

        self.check_frame = Frame(self, relief=GROOVE, border=1)
        self.check_show_rows = Checkbutton(self.check_frame, text="Show rows", variable=self.show_rows, onvalue=1, offvalue=0)
        self.check_show_row_nums = Checkbutton(self.check_frame, text="Show row numbers", variable=self.show_row_nums, onvalue=1, offvalue=0)

        self.btn_select = Button(self, text="Select workbook", command=self.on_select, relief="groove")
        self.btn_check = Button(self, text="Check workbook", command=self.on_check, relief="groove")
        self.btn_run = Button(self, text="Run", command=self.on_run, relief="groove")
        self.btn_copy = Button(self, text="Copy to clipboard", command=self.on_copy, relief="groove")
        self.btn_image = Button(self, text="Save as image file", command=self.on_save, relief="groove")
        self.btn_postscript = Button(self, text="Save as PostScript file", command=self.on_postscript, relief="groove")
        self.btn_svg = Button(self, text="Save as SVG file", command=self.on_svg, relief="groove")
        self.btn_template = Button(self, text="Download template", command=self.on_template, relief="groove")

        self.scroller = scrolledtext.ScrolledText(self, width=45, height=10, wrap=WORD,
                                                  state=DISABLED)  # defines window width

        self.pack_widgets()
        self.bind_widgets()
        self.insert_form_data()
        self.wipe_scroller()
        self.set_button_states([1, 0, 0, 0, 0, 0, 0, 1])

        self.set_select("c:/users/hayma/desktop/sample.xlsx")  # development only

    @staticmethod
    def field_validation_1(*args):
        if len(args[2]) > 5:
            return False
        elif args[2].isdigit():
            return True
        elif args[2] == "":
            return True
        else:
            return False

    @staticmethod
    def field_validation_2(*args):
        if args[2].isdigit():
            return False
        elif args[2] == "":
            return True
        else:
            return True

    @staticmethod
    def field_validation_3(*args):
        if args[2] == str(0):
            return False
        elif len(args[2]) > 3:
            return False
        elif args[2].isdigit():
            return True
        elif args[2] == "":
            return True
        else:
            return False

    def pack_widgets(self):
        self.lbl_start.grid(row=0, column=0, sticky="w", pady=(0, 0), padx=(0, 5))
        self.ent_start.grid(row=1, column=0, sticky="nsew", pady=(0, 5), padx=(0, 5))

        self.lbl_finish.grid(row=0, column=1, sticky="w", pady=(0, 0), padx=(5, 0))
        self.ent_finish.grid(row=1, column=1, sticky="nsew", pady=(0, 5), padx=(5, 0))

        self.lbl_width.grid(row=2, column=0, sticky="w", pady=(0, 0), padx=(0, 5))
        self.ent_width.grid(row=3, column=0, sticky="nsew", pady=(0, 5), padx=(0, 5))

        self.lbl_height.grid(row=2, column=1, sticky="w", pady=(0, 0), padx=(5, 0))
        self.ent_height.grid(row=3, column=1, sticky="nsew", pady=(0, 5), padx=(5, 0))

        self.lbl_num_rows.grid(row=4, column=0, sticky="w", pady=(0, 0), padx=(0, 5))
        self.ent_num_rows.grid(row=5, column=0, sticky="nsew", pady=(0, 5), padx=(0, 5))

        self.lbl_row_height.grid(row=4, column=1, sticky="w", pady=(0, 0), padx=(5, 0))
        self.ent_row_height.grid(row=5, column=1, sticky="nsew", pady=(0, 5), padx=(5, 0))

        self.check_frame.grid(row=6, column=0, columnspan=2, sticky="nsew", pady=(5, 5), padx=(0, 0))

        self.check_show_rows.grid(sticky="w", pady=(0, 0), padx=(0, 0))
        self.check_show_row_nums.grid(sticky="w", pady=(0, 0), padx=(0, 0))

        self.lbl_source.grid(row=7, column=0, columnspan=2, sticky="w")
        self.ent_filepath.grid(row=8, column=0, columnspan=2, sticky="nsew", pady=(0, 5), ipady=2)

        self.btn_select.grid(row=9, column=0, columnspan=2, sticky="nsew", pady=(0, 5))
        self.btn_check.grid(row=10, column=0, columnspan=2, sticky="nsew", pady=(0, 5))
        self.btn_run.grid(row=11, column=0, columnspan=2, sticky="nsew", pady=(0, 5))
        self.scroller.grid(row=12, column=0, columnspan=2, pady=(0, 5))
        self.btn_copy.grid(row=13, column=0, columnspan=2, sticky="nsew", pady=(0, 5))
        self.btn_image.grid(row=14, column=0, columnspan=2, sticky="nsew", pady=(0, 5))
        self.btn_postscript.grid(row=15, column=0, columnspan=2, sticky="nsew", pady=(0, 5))
        self.btn_svg.grid(row=16, column=0, columnspan=2, sticky="nsew", pady=(0, 5))
        self.btn_template.grid(row=17, column=0, columnspan=2, sticky="nsew", pady=(0, 0))  # pady 0 for last line

    def bind_widgets(self):
        self.ent_finish.bind('<FocusIn>', self.check_start)
        self.ent_start.bind('<FocusIn>', self.check_finish)

    def check_start(self, *args):
        start = self.ent_start.get_date()
        finish = self.ent_finish.get_date()
        if start > finish:
            logging.warning("Finish is less than start.")
        elif start == finish:
            logging.warning("Finish is the same as start.")
        self.refresh_scroller()

    def check_finish(self, *args):
        start = self.ent_start.get_date()
        finish = self.ent_finish.get_date()
        if start > finish:
            logging.warning("Start greater than finish.")
        elif start == finish:
            logging.warning("Start is the same as finish.")
        self.refresh_scroller()

    def get_form_data(self):
        self.settings["start"] = self.ent_start.get_date().strftime('%Y/%m/%d')
        self.settings["finish"] = self.ent_finish.get_date().strftime('%Y/%m/%d')

        width = self.ent_width.get()
        height = self.ent_height.get()
        if width:
            self.settings["width"] = width
        else:
            self.settings["width"] = 800
            logging.info("Default chart width used.")
            self.refresh_scroller()
        if height:
            self.settings["height"] = height
        else:
            self.settings["height"] = 600
            logging.info("Default chart height used.")
            self.refresh_scroller()

        num_rows = self.ent_num_rows.get()
        if num_rows:
            self.settings["num_rows"] = num_rows
        else:
            self.settings.setdefault("num_rows", 1)
            logging.info("Default number of rows used.")
            self.refresh_scroller()
        row_height = self.ent_row_height.get()
        if row_height:
            self.settings["row_height"] = row_height
        else:
            self.settings.setdefault("row_height", 10)
            logging.info("Default row height used.")
            self.refresh_scroller()

        self.settings["show_rows"] = self.show_rows.get()
        self.settings["show_row_nums"] = self.show_row_nums.get()

    def insert_form_data(self):
        self.ent_start.set_date(self.settings.get("start", datetime.date.today().strftime('%Y/%m/%d')))
        self.ent_finish.set_date(self.settings.get("finish", (datetime.date.today() + datetime.timedelta(days=10)).strftime('%Y/%m/%d')))
        self.ent_width.insert(0, self.settings.get("width", 800))
        self.ent_height.insert(0, self.settings.get("height", 600))
        self.ent_num_rows.insert(0, self.settings.get("num_rows", 1))
        self.ent_row_height.insert(0, self.settings.get("row_height", 10))
        if self.settings.get("show_rows", 1):
            self.check_show_rows.select()
        else:
            self.check_show_rows.deselect()
        if self.settings.get("show_row_nums", 1):
            self.check_show_row_nums.select()
        else:
            self.check_show_row_nums.deselect()

    def set_select(self, file_source=None):
        if file_source:
            self.file_source = file_source
            self.ent_filepath.delete(0, END)
            self.ent_filepath.insert(0, file_source)
            self.set_button_states([1, 1, 1, 0, 0, 0, 1, 1])
        else:
            self.file_source = None
            self.ent_filepath.delete(0, END)
            self.set_button_states([1, 0, 0, 0, 0, 0, 0, 1])

    def on_select(self):
        self.file_source = get_file_name(self.file_source)
        self.set_select(self.file_source)

    def on_check(self):
        workbook = load_workbook(self.file_source, data_only=True, keep_links=False)
        reference = TEMPLATE
        self.check_count += 1
        append_log(f'CHECK #{self.check_count}\n')
        check_merged_cells(workbook)
        check_sheets_exist(workbook, reference)
        check_header_rows_exist(workbook)
        check_header_rows(workbook, reference)
        check_misspelled_headers(workbook, reference)
        append_log(f'\n')
        self.refresh_scroller()

    def on_run(self):
        self.run_count += 1
        append_log(f'RUN #{self.run_count}\n')
        self.get_form_data()
        save_config_data(self.settings)
        workbook = load_workbook(self.file_source, data_only=True, keep_links=False)
        items = Loader(workbook).items
        items = Cleaner(items).items
        items = Processor(items).items
        if self.view:
            self.view.destroy()
        self.view = View(parent=self.parent, data=items)  # App is the parent
        self.set_button_states([1, 1, 1, 1, 1, 1, 1, 1])
        append_log(f'Run {self.run_count} complete.\n\n')
        self.refresh_scroller()

    def on_copy(self):
        copy_to_clipboard(self.view.image)
        self.refresh_scroller()

    def on_save(self):
        save_image(self.view.image)
        self.refresh_scroller()

    def on_postscript(self):
        save_postscript(self.view.image)
        self.refresh_scroller()

    def on_svg(self):
        self.get_form_data()
        save_config_data(self.settings)
        workbook = load_workbook(self.file_source, data_only=True, keep_links=False)
        items = Loader(workbook).items
        items = Cleaner(items).items
        items = Processor(items).items
        items = Generator(items).items
        settings = Settings()
        elements = ''.join([item.element for item in items])
        svg = f'<svg width="{settings.width}" height="{settings.height}" id="chart" overflow="auto">{elements}</svg>'.encode(encoding='utf-8')
        file = io.BytesIO(svg)
        file.name = 'chart.svg'
        rlg = svg2rlg(file)
        renderPM.drawToFile(rlg, "chart.png", fmt="PNG")

    def on_template(self):
        workbook = create_template(TEMPLATE)
        workbook = populate_template(workbook, SAMPLE)
        workbook = reformat_dates(workbook)
        export_workbook(workbook)
        self.refresh_scroller()

    def set_button_states(self, states=None):
        buttons = [self.btn_select, self.btn_check, self.btn_run, self.btn_copy, self.btn_image, self.btn_postscript, self.btn_svg, self.btn_template]
        if not states:
            states = [1, 0, 0, 0, 0, 0, 0, 1]
        states = [NORMAL if x == 1 else DISABLED for x in states]  # swaps values for variables
        for button, state in zip(buttons, states):
            button.config(state=state)

    def wipe_scroller(self):
        self.scroller.config(state=NORMAL)
        self.scroller.delete('1.0', END)  # from line '1' (entry equivalent is from char 0)
        self.scroller.config(state=DISABLED)

    def refresh_scroller(self):
        self.wipe_scroller()
        log = get_log()
        self.scroller.configure(state=NORMAL)  # writable
        self.scroller.insert(END, log)
        self.scroller.configure(state=DISABLED)  # readable
        self.scroller.see(END)


class View(Toplevel):
    """This is the parent of Canvas.  It determines image size (i.e. the Canvas does not do so)."""
    def __init__(self, parent, data):
        super(View, self).__init__(parent)

        self.win_x = int(self.winfo_screenwidth() * 0.4)
        self.win_y = int(self.winfo_screenheight() * 0.1)
        self.geometry(f'+{self.win_x}+{self.win_y}')  # w, h, x, y
        self.resizable(False, False)
        self.title("Gantt Page")
        self.wm_iconbitmap(get_path("favicon.ico"))
        self.parent = parent
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.image = Painter(self, data)  # View is the parent

    def on_close(self):
        self.parent.controls.set_button_states([1, 1, 1, 0, 0, 0, 1, 1])
        self.destroy()
